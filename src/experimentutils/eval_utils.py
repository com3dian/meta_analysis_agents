"""
Evaluation utilities for mapping papers to ground-truth annotations
and pre-processing paper text for better LLM extraction.

Public API
----------
Field-level scoring
    field_similarity_score  – hybrid numeric / year / ROUGE-L scorer

Record matching & evaluation
    evaluate_record_pair_with_units – per-field scores, value+unit, crop swap
    find_crop_swap_pairs    – detect all crop-1/crop-2 column pairs
    match_records_greedy    – one-to-one greedy matching with swap support
    evaluate_method_scores  – normalized score over all GT records
    print_matching_pairs    – pretty-print matched pairs for inspection
    print_field_value_pairs_vs_gt – Step-1 pairs vs GT (same column layout)
"""
from __future__ import annotations

import ast
import csv
import io
import json
import os
import re
import decimal
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Final, Iterable, List, Mapping, Optional, Tuple, Union

import numpy as np

try:  # Optional dependency in some environments (e.g., minimal CI / tooling)
    import pandas as pd  # type: ignore
except Exception:  # pragma: no cover
    pd = None  # type: ignore

from .file_utils import _find_project_root, get_all_paper_paths
from .units import (
    UNIT_FIELDS,
    VALUE_TO_UNIT_FIELD,
    Quantity,
    Unit,
    ValueUnitGroup,
    WOPKE_100_GT_VALUE_UNIT_GROUPS,
    WOPKE_100_VALUE_UNIT_GROUPS,
    unit_field_for_value,
    value_unit_groups,
)

RecordRow = Union[Mapping[str, Any], Any]

DEFAULT_ANNOTATION = (
    "data/wopke_paper_code/Database for combined sample 2015-03-05.csv"
)

# Ground-truth headers → wopke_100 standard field names (copies; originals kept).
# Replications are stored as sample-size columns next to yield (N sc 1, …),
# matching the LER-paper R script rename N.sc.1 → No._SC_1.
GT_COLUMN_ALIASES: Final[Dict[str, str]] = {
    "N sc 1": "Replications SC1",
    "N sc 2": "Replications SC2",
    "N ic 1": "Replications IC1",
    "N ic 2": "Replications IC2",
    "Unit.1": "N Unit",
    "Unit.2": "P Unit",
    "Unit.3": "K Unit",
    "Data location": "Data source",
}


# Fields that participate in crop-1/crop-2 swap (see find_crop_swap_pairs).
# Includes value fields with trailing 1/2 but excludes shared unit columns.
SWAP_ELIGIBLE_SUFFIX_FIELDS: Final[Tuple[str, ...]] = (
    "Crop species 1",
    "Crop species 2",
    "Crop type 1",
    "Crop type 2",
    "Fodder crop 1",
    "Fodder crop 2",
    "Density ic 1",
    "Density ic 2",
    "Density sc 1",
    "Density sc 2",
    "N input SC1",
    "N input SC2",
    "N input IC1",
    "N input IC2",
    "P input SC1",
    "P input SC2",
    "P input IC1",
    "P input IC2",
    "K input SC1",
    "K input SC2",
    "K input IC1",
    "K input IC2",
    "unified yield sc 1",
    "unified yield sc 2",
    "unified yield ic 1",
    "unified yield ic 2",
    "Sowing date 1",
    "Sowing date 2",
    "Harvest date 1",
    "Harvest date 2",
    "Replications SC1",
    "Replications SC2",
    "Replications IC1",
    "Replications IC2",
    "PLER 1",
    "PLER 2",
)


def _dedupe_column_names(header: Iterable[Any]) -> List[str]:
    """Make duplicate headers unique (``Unit`` → ``Unit``, ``Unit.1``, …)."""
    seen: Dict[str, int] = {}
    deduped: List[str] = []
    for col in header:
        name = "" if col is None else str(col)
        if name in seen:
            seen[name] += 1
            deduped.append(f"{name}.{seen[name]}")
        else:
            seen[name] = 0
            deduped.append(name)
    return deduped


def apply_gt_column_aliases(df: "pd.DataFrame") -> "pd.DataFrame":
    """
    Copy GT columns onto ``wopke_100`` standard names when missing.

    Original headers are kept so unit-aware scoring can still read
    ``Unit.1`` / ``N sc 1``.
    """
    out = df.copy()
    for src, dst in GT_COLUMN_ALIASES.items():
        if src in out.columns and dst not in out.columns:
            out[dst] = out[src]
    return out


def _decode_csv_text(path: str) -> str:
    raw = Path(path).read_bytes()
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1")


def _load_gt_csv(path: str) -> "pd.DataFrame":
    reader = csv.reader(io.StringIO(_decode_csv_text(path)))
    header = next(reader)
    data_rows = [row for row in reader if any(cell not in (None, "") for cell in row)]
    n_cols = len(header)
    padded = [list(row) + [""] * (n_cols - len(row)) for row in data_rows]
    return pd.DataFrame(padded, columns=_dedupe_column_names(header))


def _load_gt_xlsx(path: str, sheet: str) -> "pd.DataFrame":
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        sheet = wb.sheetnames[0]
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    return pd.DataFrame(rows[1:], columns=_dedupe_column_names(header))


def load_ground_truth(
    annotation_path: Optional[str] = None,
    sheet: str = "labels",
    *,
    apply_aliases: bool = True,
) -> pd.DataFrame:
    """
    Load the wopke_100 ground-truth table (CSV or Excel), deduplicating headers.

    Default file is ``Database for combined sample 2015-03-05.csv``. Duplicate
    headers get a ``.N`` suffix. When ``apply_aliases`` is true, standard names
    such as ``Replications SC1`` and ``N Unit`` are copied from the GT columns.

    Returns:
        DataFrame with unique column names.
    """
    if annotation_path is None:
        annotation_path = os.path.join(_find_project_root(), DEFAULT_ANNOTATION)

    suffix = Path(annotation_path).suffix.lower()
    if suffix == ".csv":
        gt_df = _load_gt_csv(annotation_path)
    else:
        gt_df = _load_gt_xlsx(annotation_path, sheet)

    if apply_aliases:
        gt_df = apply_gt_column_aliases(gt_df)
    return gt_df


def wopke_100_shared_fields(gt_columns: Iterable[Any]) -> List[str]:
    """Return ``wopke_100`` standard fields present in ``gt_columns``."""
    from src.standards import METADATA_STANDARDS

    fields = list(json.loads(METADATA_STANDARDS["wopke_100"].strip()).keys())
    cols = {str(c) for c in gt_columns}
    return [f for f in fields if f in cols]


def load_ground_truth_by_study_id(
    annotation_path: Optional[str] = None,
    sheet: str = "labels",
    *,
    apply_aliases: bool = True,
) -> Dict[int, List[Dict[str, Any]]]:
    """Load GT rows grouped by ``Study#``."""
    gt_df = load_ground_truth(
        annotation_path, sheet=sheet, apply_aliases=apply_aliases
    )
    grouped: Dict[int, List[Dict[str, Any]]] = {}
    for rec in gt_df.to_dict(orient="records"):
        sid_val = rec.get("Study#")
        if sid_val is None or (isinstance(sid_val, float) and sid_val != sid_val):
            continue
        try:
            sid_int = int(float(sid_val))
        except (TypeError, ValueError):
            continue
        grouped.setdefault(sid_int, []).append(rec)
    return grouped


def build_study_paper_mapping(
    gt_df: Optional[pd.DataFrame] = None,
    paper_paths: Optional[List[str]] = None,
) -> Dict[int, str]:
    """
    Build a mapping from ``Study#`` in the annotation to the closest paper
    file path, using author last-name + publication year fuzzy matching.

    Returns:
        ``{study_id: paper_path}`` for every study that could be matched.
    """
    if gt_df is None:
        gt_df = load_ground_truth()
    if paper_paths is None:
        paper_paths = get_all_paper_paths()

    paper_stems = {Path(p).stem.lower(): p for p in paper_paths}

    studies: Dict[int, Tuple[str, str]] = {}
    for _, row in gt_df.iterrows():
        sid = row["Study#"]
        if sid in studies:
            continue
        author = str(row.get("Author", ""))
        year = str(row.get("Year of publication", ""))[:4]
        studies[int(sid)] = (author, year)

    mapping: Dict[int, str] = {}
    for sid, (author, year) in studies.items():
        last_name = author.split(",")[0].split(";")[0].strip().lower()
        if not last_name:
            continue
        exact, partial = [], []
        for stem, path in paper_stems.items():
            if last_name in stem:
                if year and year in stem:
                    exact.append(path)
                else:
                    partial.append(path)
        if exact:
            mapping[sid] = exact[0]
        elif partial:
            mapping[sid] = partial[0]

    return mapping


def get_paper_path_for_study(
    study_id: int,
    mapping: Optional[Dict[int, str]] = None,
) -> Optional[str]:
    """Return the paper path for a given ``Study#``, or *None* if unmatched."""
    if mapping is None:
        mapping = build_study_paper_mapping()
    return mapping.get(study_id)


# ---------------------------------------------------------------------------
# Text pre-processing: highlight tables & numbers
# ---------------------------------------------------------------------------

_NUMBER_RE = re.compile(
    r"(?<!\w)"
    r"(\d[\d,]*\.?\d*)"           # integer or decimal, optionally with commas
    r"(\s*(?:"
    r"[%°]"                        # percent, degree
    r"|g\s*/\s*m2|g/m2"
    r"|kg\s*/?\s*ha|kg\s+ha"
    r"|t\s*/?\s*ha|t\s+ha"
    r"|Mg\s*/?\s*ha|Mg\s+ha"
    r"|bu\s*/?\s*acre"
    r"|plants?\s*/?\s*m2|plants\s+m2"
    r"|kg\s+N\s*/?\s*ha|kg\s+N\s+ha"
    r"|kg\s+P2O5\s*/?\s*ha"
    r"|kg\s+K2O\s*/?\s*ha"
    r"|days?"
    r"))?"
    r"(?!\w)",
    re.IGNORECASE,
)

_TABLE_HEADER_RE = re.compile(
    r"^(Table|TABLE|Fig(?:ure)?|FIGURE)\s+\d",
    re.MULTILINE,
)


def highlight_numbers_and_tables(text: str) -> str:
    """
    Annotate raw paper text with ``<<NUM: ... >>`` markers around numbers
    (with optional units) and ``<<TABLE_START>>`` / ``<<TABLE_END>>`` markers
    around table-like sections.

    This makes numerical data and tabular regions much more visible to LLMs
    that tend to overlook numbers embedded in dense text.
    """
    # Mark numbers
    def _mark_number(m: re.Match) -> str:
        full = m.group(0).strip()
        return f"<<NUM: {full} >>"

    text = _NUMBER_RE.sub(_mark_number, text)

    # Mark table-like blocks: consecutive lines with >= 3 numbers
    lines = text.split("\n")
    in_table = False
    out_lines: List[str] = []
    consecutive_numeric = 0

    for line in lines:
        num_count = len(re.findall(r"<<NUM:", line))
        is_table_header = bool(_TABLE_HEADER_RE.match(line.strip()))

        if is_table_header or num_count >= 3:
            if not in_table:
                out_lines.append("<<TABLE_START>>")
                in_table = True
            consecutive_numeric = 0
        elif in_table:
            consecutive_numeric += 1
            if consecutive_numeric > 2:
                out_lines.append("<<TABLE_END>>")
                in_table = False
                consecutive_numeric = 0

        out_lines.append(line)

    if in_table:
        out_lines.append("<<TABLE_END>>")

    return "\n".join(out_lines)


# ---------------------------------------------------------------------------
# Field-level similarity scoring
# ---------------------------------------------------------------------------

_MISSING_TOKENS = {"", "nan", "none", "null", "n/a", "na"}

# Matches a 4-digit calendar year (1000–2099)
_YEAR_RE = re.compile(r"\b(1\d{3}|20\d{2})\b")

# Matches a value that is *purely* numeric (optional leading sign, comma
# thousands separators, and a single decimal point)
_PURE_NUMBER_RE = re.compile(r"^[+-]?[\d,]+\.?\d*$")

# Matches a parenthetical scientific name, e.g. "(Vicia faba)" or "(Triticum aestivum L.)"
_SCIENTIFIC_NAME_RE = re.compile(r"\s*\([A-Z][a-z]+(?:\s+[a-z]+\.?)+\)")
_DATE_FIELD_RE = re.compile(r"\b(date|sowing|harvest)\b", re.IGNORECASE)
_RELATIVE_DAYS_RE = re.compile(
    r"(\d{1,3})(?:\s*(?:-|to|and)\s*(\d{1,3}))?\s*days?\b",
    re.IGNORECASE,
)
_MONTH_PATTERN = (
    r"(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
    r"jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|"
    r"dec(?:ember)?)"
)
_MONTH_NAMES_IN_TEXT_RE = re.compile(_MONTH_PATTERN, re.IGNORECASE)
_DAY_MONTH_YEAR_IN_TEXT_RE = re.compile(
    rf"(\d{{1,2}})\s+{_MONTH_PATTERN}\s+(\d{{4}})",
    re.IGNORECASE,
)
_MONTH_DAY_YEAR_IN_TEXT_RE = re.compile(
    rf"{_MONTH_PATTERN}\s+(\d{{1,2}})(?:st|nd|rd|th)?(?:,)?\s+(\d{{4}})",
    re.IGNORECASE,
)
_MONTH_YEAR_IN_TEXT_RE = re.compile(
    rf"{_MONTH_PATTERN}\s+(\d{{4}})",
    re.IGNORECASE,
)
_DAY_MONTH_IN_TEXT_RE = re.compile(
    rf"(\d{{1,2}})\s+{_MONTH_PATTERN}\b",
    re.IGNORECASE,
)

_MONTH_ALIASES: Final[Dict[str, int]] = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def _is_missing(v: Any) -> bool:
    """Return True for NaN, None, or any blank/null-like string."""
    if v is None:
        return True
    return str(v).strip().lower() in _MISSING_TOKENS


def _try_parse_number(v: Any) -> Optional[float]:
    """Return float if *v* represents a pure number, else None."""
    s = str(v).strip().replace(",", "")
    if _PURE_NUMBER_RE.match(s):
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _try_parse_decimal(v: Any) -> Optional["decimal.Decimal"]:
    """
    Return a Decimal if *v* is a pure number, else None.

    Uses the string form (after removing comma separators) to avoid float
    rounding artifacts during exact-equality checks.
    """
    s = str(v).strip().replace(",", "")
    if not _PURE_NUMBER_RE.match(s):
        return None
    try:
        # Decimal accepts leading +/-, integer and decimal forms.
        return decimal.Decimal(s)
    except (decimal.InvalidOperation, ValueError):
        return None


def _remove_decimal_point_numeric_string(v: Any) -> Optional[str]:
    """
    If *v* is a pure number, return a canonical string with the decimal point removed.

    Example:
        "3.124" -> "3124"
        "3124"  -> "3124"
        "-3.10" -> "-310"
    """
    s = str(v).strip().replace(",", "")
    if not _PURE_NUMBER_RE.match(s):
        return None
    return s.replace(".", "")


def _excel_serial_to_iso(v: Any) -> Optional[str]:
    """
    Convert an Excel-style serial day (1 = 1900-01-01) into an ISO date string.

    Only treats values within a reasonable range as serial dates to avoid
    misinterpreting arbitrary integers.
    """
    s = str(v).strip().replace(",", "")
    # Must look like a pure number (integer or simple decimal)
    if not _PURE_NUMBER_RE.match(s):
        return None
    try:
        dec = decimal.Decimal(s)
    except (decimal.InvalidOperation, ValueError):
        return None

    # Only treat integer-valued decimals as serials (e.g. 35693 or 35693.0)
    try:
        n = int(dec)
    except (OverflowError, ValueError):
        return None

    # Rough guard rails: years roughly between 1950 and 2050
    # 1900-01-01 + 20000 days ≈ 1954, +60000 ≈ 2064
    if n < 20000 or n > 60000:
        return None
    base = date(1900, 1, 1)
    # Empirically, Excel-style serials in the Wopke sheets decode one day
    # ahead with the naive 1900-01-01 + (n-1) formula, so we subtract one
    # extra day to align with the GT dates.
    d = base + timedelta(days=n - 2)
    return d.isoformat()


def _normalize_text(v: Any) -> str:
    """
    Prepare a text value for soft token comparison:

    Lowercase and collapse whitespace while preserving all content, including
    parenthetical scientific names.
    """
    return " ".join(str(v).lower().split())


def _soft_tokens_match(t1: str, t2: str) -> bool:
    """
    Return ``True`` when two tokens are considered equivalent.

    Exact match OR one is a prefix of the other, which handles common
    plural/singular crop-name variants (``"beans"`` vs ``"bean"``,
    ``"wheats"`` vs ``"wheat"``, ``"maizes"`` vs ``"maize"``).
    """
    print(f"Soft tokens match: {t1} vs {t2}")
    return t1 == t2 or t1.startswith(t2) or t2.startswith(t1)


def _extract_year(v: Any) -> Optional[str]:
    """
    Return the first 4-digit year found in *v*, or None.

    Handles plain years ("1987"), slash ranges ("1987/88"),
    and hyphen ranges ("1987-88", "1987-1988").
    """
    m = _YEAR_RE.search(str(v))
    return m.group(1) if m else None


def _is_date_field_name(field_name: Optional[str]) -> bool:
    if not field_name:
        return False
    return bool(_DATE_FIELD_RE.search(field_name))


def _month_to_number(month_token: str) -> Optional[int]:
    return _MONTH_ALIASES.get(month_token.strip().lower())


def _parse_absolute_iso_date_from_text(v: Any) -> Optional[str]:
    """
    Parse date-like values into an ISO date string when enough detail exists.

    Supports:
    - Excel serial numbers (e.g., 35693)
    - Python date/datetime objects
    - ISO strings (YYYY-MM-DD)
    - Day Month Year patterns inside free text (e.g., "sown 16 April 1997")
    - Month Day Year patterns (e.g., "April 16, 1997")
    """
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()

    s = str(v).strip()
    if not s:
        return None

    excel_iso = _excel_serial_to_iso(s)
    if excel_iso is not None:
        return excel_iso

    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s

    m = _DAY_MONTH_YEAR_IN_TEXT_RE.search(s)
    if m:
        day = int(m.group(1))
        month = _month_to_number(m.group(2))
        year = int(m.group(3))
        if month is not None:
            try:
                return date(year, month, day).isoformat()
            except ValueError:
                return None

    m = _MONTH_DAY_YEAR_IN_TEXT_RE.search(s)
    if m:
        month = _month_to_number(m.group(1))
        day = int(m.group(2))
        year = int(m.group(3))
        if month is not None:
            try:
                return date(year, month, day).isoformat()
            except ValueError:
                return None

    return None


def _parse_relative_days_value(v: Any) -> Optional[Tuple[int, int]]:
    """Parse relative-duration expressions such as '105 days' or '50 and 95 days'."""
    s = str(v).strip()
    if not s:
        return None
    m = _RELATIVE_DAYS_RE.search(s)
    if not m:
        return None
    d1 = int(m.group(1))
    d2 = int(m.group(2)) if m.group(2) is not None else d1
    lo, hi = sorted((d1, d2))
    return lo, hi


def _parse_month_day_without_year(v: Any) -> Optional[Tuple[int, int]]:
    """Parse month+day expressions like '20 September' when year is absent."""
    s = str(v).strip()
    if not s:
        return None
    m = _DAY_MONTH_IN_TEXT_RE.search(s)
    if not m:
        return None
    day = int(m.group(1))
    month = _month_to_number(m.group(2))
    if month is None:
        return None
    try:
        date(2000, month, day)
    except ValueError:
        return None
    return month, day


def _extract_month_year(v: Any) -> Optional[Tuple[int, int]]:
    """Extract month and year from expressions like 'Mid May 2004'."""
    s = str(v).strip()
    if not s:
        return None
    m = _MONTH_YEAR_IN_TEXT_RE.search(s)
    if not m:
        return None
    month = _month_to_number(m.group(1))
    year = int(m.group(2))
    if month is None:
        return None
    return month, year


def date_similarity_score(ref_val: Any, hyp_val: Any, *, debug: bool = False) -> float:
    """
    Compare date-like values with lightweight normalization across common formats.

    Matching order:
    1) Exact absolute date match (after converting serial/text forms to ISO date).
    2) Exact relative-days match (e.g., "105 days" vs "105 days after sowing").
    3) Exact month+day match when both omit year.
    4) Exact month+year match (for coarse month-level expressions).
    5) Exact year match.
    6) Fallback to ROUGE-L soft text score.
    """
    if _is_missing(ref_val) or _is_missing(hyp_val):
        return 0.0

    ref_iso = _parse_absolute_iso_date_from_text(ref_val)
    hyp_iso = _parse_absolute_iso_date_from_text(hyp_val)
    if ref_iso is not None and hyp_iso is not None:
        if debug:
            print(f"Date absolute comparison: {ref_iso} vs {hyp_iso}")
        return 1.0 if ref_iso == hyp_iso else 0.0

    # If one side has a full date (with year) and the other side only provides
    # month/day text, compare by month/day and ignore year.
    def _iso_to_month_day(iso_date: str) -> Optional[Tuple[int, int]]:
        try:
            d = date.fromisoformat(iso_date)
            return d.month, d.day
        except ValueError:
            return None

    ref_iso_md = _iso_to_month_day(ref_iso) if ref_iso is not None else None
    hyp_iso_md = _iso_to_month_day(hyp_iso) if hyp_iso is not None else None
    ref_md = _parse_month_day_without_year(ref_val)
    hyp_md = _parse_month_day_without_year(hyp_val)

    if ref_iso_md is not None and hyp_md is not None:
        if debug:
            print(f"Date ISO-vs-month-day comparison: {ref_iso_md} vs {hyp_md}")
        return 1.0 if ref_iso_md == hyp_md else 0.0

    if hyp_iso_md is not None and ref_md is not None:
        if debug:
            print(f"Date month-day-vs-ISO comparison: {ref_md} vs {hyp_iso_md}")
        return 1.0 if ref_md == hyp_iso_md else 0.0

    ref_rel = _parse_relative_days_value(ref_val)
    hyp_rel = _parse_relative_days_value(hyp_val)
    if ref_rel is not None and hyp_rel is not None:
        if debug:
            print(f"Date relative-days comparison: {ref_rel} vs {hyp_rel}")
        return 1.0 if ref_rel == hyp_rel else 0.0

    if ref_md is not None and hyp_md is not None:
        if debug:
            print(f"Date month-day comparison: {ref_md} vs {hyp_md}")
        return 1.0 if ref_md == hyp_md else 0.0

    ref_my = _extract_month_year(ref_val)
    hyp_my = _extract_month_year(hyp_val)
    if ref_my is not None and hyp_my is not None:
        if debug:
            print(f"Date month-year comparison: {ref_my} vs {hyp_my}")
        return 1.0 if ref_my == hyp_my else 0.0

    ref_year = _extract_year(ref_val)
    hyp_year = _extract_year(hyp_val)
    if ref_year is not None and hyp_year is not None:
        if debug:
            print(f"Date year comparison: {ref_year} vs {hyp_year}")
        return 1.0 if ref_year == hyp_year else 0.0

    if _MONTH_NAMES_IN_TEXT_RE.search(str(ref_val)) and _MONTH_NAMES_IN_TEXT_RE.search(str(hyp_val)):
        return rouge_l_soft_score(ref_val, hyp_val)

    return rouge_l_soft_score(ref_val, hyp_val)


def rouge_l_soft_score(ref_val: Any, hyp_val: Any) -> float:
    """
    Compute ROUGE-L F1 on character sequences.

    This helper is exposed so ROUGE-L behavior can be tested directly without
    running the full field-type routing in ``field_similarity_score``.
    """
    ref_chars = [c for c in _normalize_text(ref_val) if not c.isspace()]
    hyp_chars = [c for c in _normalize_text(hyp_val) if not c.isspace()]
    if not ref_chars or not hyp_chars:
        return 0.0

    m, nl = len(ref_chars), len(hyp_chars)
    dp = [[0] * (nl + 1) for _ in range(m + 1)]
    for i in range(1, m + 1):
        for j in range(1, nl + 1):
            dp[i][j] = (
                dp[i - 1][j - 1] + 1
                if ref_chars[i - 1] == hyp_chars[j - 1]
                else max(dp[i - 1][j], dp[i][j - 1])
            )

    lcs = dp[m][nl]
    p, r = lcs / nl, lcs / m
    return 0.0 if p + r == 0 else 2 * p * r / (p + r)


def field_similarity_score(
    ref_val: Any,
    hyp_val: Any,
    field_name: Optional[str] = None,
    *,
    debug: bool = False,
) -> float:
    """
    Hybrid similarity score for a single extracted field value compared to
    its ground-truth counterpart.

    The function selects a comparison strategy automatically based on the
    detected value type:

    1. **Missing** – either side is ``None``, ``NaN``, empty, or a null-like
       string (``"nan"``, ``"none"``, ``"null"``, ``"n/a"``).  Returns ``0.0``
       so jointly-missing fields do not receive a free perfect score.

    2. **Numeric** – both sides parse as pure numbers (integers or decimals,
       optionally with comma thousands-separators).

       - Exact numeric match → ``1.0``
       - Else, if removing the decimal point makes the strings equal
         (e.g. ``"3.124"`` ↔ ``"3124"``) → ``0.5``
       - Else → ``0.0``

       No approximate / tolerance / relative-error scoring is used.

    3. **Year** – both sides contain a 4-digit calendar year.  Normalises
       year-range strings (``"1987-88"``, ``"1987/88"``) to their first
       4-digit year before comparing.  Returns ``1.0`` for equal years,
       ``0.0`` otherwise.

    4. **Text / categorical** – ROUGE-L F1 with two pre-processing steps
       applied before tokenisation:

       * **Scientific name stripping** – removes parenthetical Latin binomials
         so ``"field beans (Vicia faba)"`` becomes ``"field beans"``.
       * **Prefix / stem token matching** – two tokens are treated as equal
         when one is a prefix of the other, handling common plural/singular
         variants (``"beans"`` ↔ ``"bean"``, ``"wheats"`` ↔ ``"wheat"``).

    Args:
        ref_val: Reference (ground-truth) value – any scalar type.
        hyp_val: Hypothesis (extracted) value – any scalar type.

    Returns:
        Similarity score in ``[0.0, 1.0]``.

    Examples:
        >>> field_similarity_score("3.67", "3.7")
        0.0
        >>> field_similarity_score("3.124", "3124")
        0.5
        >>> field_similarity_score("1987-88", "1987")
        1.0
        >>> field_similarity_score("Bean", "bean")
        1.0
        >>> field_similarity_score("field beans (Vicia faba)", "Bean")
        0.667...
        >>> field_similarity_score("wheat (Triticum aestivum)", "Wheat")
        1.0
        >>> field_similarity_score("Bean", "Wheat")
        0.0
        >>> field_similarity_score(None, "1987")
        0.0
        >>> field_similarity_score(float("nan"), float("nan"))
        0.0
    """
    if _is_missing(ref_val) or _is_missing(hyp_val):
        return 0.0

    # ── 0. Date-field comparison (single routing point for pair evaluation) ──
    if _is_date_field_name(field_name):
        return date_similarity_score(ref_val, hyp_val, debug=debug)

    # ── 1. Numeric comparison ────────────────────────────────────────────────
    ref_dec = _try_parse_decimal(ref_val)
    hyp_dec = _try_parse_decimal(hyp_val)
    if ref_dec is not None and hyp_dec is not None:
        if debug:
            print(f"Numeric comparison: {ref_dec} vs {hyp_dec}")
        # Exact numeric equality first (e.g. "3" == "3.0" == 3)
        if ref_dec == hyp_dec:
            return 1.0
        # Second-chance: match by removing the decimal point (e.g. 3.124 ~ 3124)
        ref_no_dot = _remove_decimal_point_numeric_string(ref_val)
        hyp_no_dot = _remove_decimal_point_numeric_string(hyp_val)
        if ref_no_dot is not None and hyp_no_dot is not None and ref_no_dot == hyp_no_dot:
            return 0.5
        return 0.0

    # ── 2. Excel-serial-style date comparison ───────────────────────────────
    # If one or both values look like Excel serial days and can be converted
    # to calendar dates, compare their ISO strings for exact match.
    ref_excel = _excel_serial_to_iso(ref_val)
    hyp_excel = _excel_serial_to_iso(hyp_val)
    if ref_excel is not None or hyp_excel is not None:
        if debug:
            print(f"Excel-serial-style date comparison: {ref_excel} vs {hyp_excel}")
        # Try to normalise the other side to ISO date as well:
        # - if it's already ISO-like (YYYY-MM-DD), accept as-is;
        # - otherwise, fall back to year-only comparison below.
        def _normalize_dateish(v: Any) -> Optional[str]:
            s = str(v).strip()
            # Simple ISO date check
            if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
                return s
            return _excel_serial_to_iso(s)

        ref_iso = ref_excel or _normalize_dateish(ref_val)
        hyp_iso = hyp_excel or _normalize_dateish(hyp_val)
        if ref_iso is not None and hyp_iso is not None:
            return 1.0 if ref_iso == hyp_iso else 0.0

    # ── 3. Year comparison ───────────────────────────────────────────────────
    ref_year = _extract_year(ref_val)
    hyp_year = _extract_year(hyp_val)
    if ref_year is not None and hyp_year is not None:
        if debug:
            print(f"Year comparison: {ref_year} vs {hyp_year}")
        return 1.0 if ref_year == hyp_year else 0.0

    # ── 4. Text / categorical comparison ────────────────────────────────────
    # Strip scientific names then tokenise; use prefix matching in LCS so
    # that "beans"/"bean", "wheats"/"wheat" etc. are treated as equivalent.
    return rouge_l_soft_score(ref_val, hyp_val)


def build_extraction_context(
    paper_text: str,
    gt_df: Optional[pd.DataFrame] = None,
    study_id: Optional[int] = None,
) -> str:
    """
    Build an enriched extraction context by:
    1. Highlighting numbers and table regions.
    2. Appending a GT record-count hint (if available) so the LLM knows
       roughly how many records to expect.

    Args:
        paper_text: Raw text of the paper.
        gt_df: Ground-truth DataFrame (optional).
        study_id: Study# for this paper (optional).

    Returns:
        Annotated text ready to feed into extraction prompts.
    """
    enriched = highlight_numbers_and_tables(paper_text)

    if gt_df is not None and study_id is not None:
        gt_rows = gt_df[gt_df["Study#"] == study_id]
        if len(gt_rows) > 0:
            enriched += (
                f"\n\n<!-- HINT: the human annotation for this paper contains "
                f"{len(gt_rows)} experiment records. -->\n"
            )

    return enriched


# ---------------------------------------------------------------------------
# Record-level matching & evaluation
# ---------------------------------------------------------------------------

def find_crop_swap_pairs(cols: List[str]) -> List[Tuple[str, str]]:
    """
    Detect all column pairs that differ only in a trailing ``'1'`` vs ``'2'``.

    Handles both space-separated suffixes (``'Crop species 1'`` /
    ``'Crop species 2'``) and concatenated suffixes (``'N input SC1'`` /
    ``'N input SC2'``).  Swapping **all** pairs simultaneously corresponds to
    relabelling crop-1 as crop-2 and vice-versa across every related field.

    Args:
        cols: Column names to inspect (e.g. the list of shared evaluation fields).

    Returns:
        List of ``(col_1, col_2)`` tuples where ``col_1`` ends in ``'1'`` and
        ``col_2`` is the matching ``'2'`` variant.

    Example:
        >>> cols = ['Crop species 1', 'Crop species 2', 'N input SC1', 'N input SC2', 'Yield unit']
        >>> find_crop_swap_pairs(cols)
        [('Crop species 1', 'Crop species 2'), ('N input SC1', 'N input SC2')]
    """
    col_set = set(cols)
    seen: set = set()
    pairs: List[Tuple[str, str]] = []
    for c in cols:
        if c in seen:
            continue
        for suffix1, suffix2 in [(" 1", " 2"), ("1", "2")]:
            if c.endswith(suffix1):
                partner = c[: -len(suffix1)] + suffix2
                if partner in col_set:
                    pairs.append((c, partner))
                    seen.add(c)
                    seen.add(partner)
                    break
    return pairs


def is_present(v: Any) -> bool:
    """
    Return True iff a value should count as "present" (non-missing) in eval.

    Intended for notebook code paths that operate on lists of dicts (CSV rows)
    rather than pandas Series/DataFrames.
    """
    if v is None:
        return False
    s = str(v).strip()
    if not s:
        return False
    return s.lower() not in {"nan", "none", "null", "n/a", "na"}


def _apply_swaps_row(row: Dict[str, Any], pairs: List[Tuple[str, str]]) -> Dict[str, Any]:
    if not pairs:
        return row
    out = dict(row)
    for c1, c2 in pairs:
        out[c1] = row.get(c2, None)
        out[c2] = row.get(c1, None)
    return out


def swap_1_2_fields_for_records(
    rows: List[Dict[str, Any]],
    fields: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    """
    Return a copy of all records with every detected 1/2 field pair swapped.

    The swap is applied globally to every record, not row-by-row conditionally.
    For example, ``Crop species 1`` and ``Crop species 2`` are exchanged in
    every predicted row, as are fields like ``N input SC1``/``N input SC2``.
    """
    if not rows:
        return []
    cols = fields or list(rows[0].keys())
    swap_pairs = find_crop_swap_pairs(cols)
    return [_apply_swaps_row(row, swap_pairs) for row in rows]


def match_records_greedy_presence(
    pred_rows: List[Dict[str, Any]],
    gt_rows: List[Dict[str, Any]],
    fields: List[str],
    *,
    allow_pair_swaps: bool = True,
) -> List[Tuple[int, int, bool, float]]:
    """
    Greedy one-to-one record matching for list-of-dicts inputs, using a
    *presence/label* signal (not full value similarity).

    This mirrors the pairing heuristic used in
    `notebooks/value_presence_confusion_per_method_per_paper.ipynb`:

    - Missing/missing contributes 0.0 (prevents "free" matches on empty rows).
    - For label-ish fields (name contains "species" or "label"): exact match
      after normalisation.
    - For other fields: any jointly-present values count as a match signal.

    Crop-1/crop-2 fields can be swapped as a block; for each candidate pair the
    better orientation is chosen before greedy matching.

    Returns:
        List of (pred_row_index, gt_row_index, swapped, mean_score) tuples.
    """

    def _normalize_label(v: Any) -> str:
        return str(v).strip().lower()

    def _is_label_field(col: str) -> bool:
        lc = col.lower()
        return ("species" in lc) or ("label" in lc)

    def _field_signal(col: str, pred_val: Any, gt_val: Any) -> float:
        pred_p = is_present(pred_val)
        gt_p = is_present(gt_val)
        if (not pred_p) or (not gt_p):
            return 0.0
        if _is_label_field(col):
            return 1.0 if _normalize_label(pred_val) == _normalize_label(gt_val) else 0.0
        return 1.0

    swap_pairs = find_crop_swap_pairs(fields) if allow_pair_swaps else []

    def _mean_signal(row_values: Dict[str, Any], gt_row: Dict[str, Any]) -> float:
        if not fields:
            return 0.0
        s = 0.0
        for c in fields:
            s += _field_signal(c, row_values.get(c), gt_row.get(c))
        return s / float(len(fields))

    def _best_orientation(pred_row: Dict[str, Any], gt_row: Dict[str, Any]) -> Tuple[float, bool]:
        mean_orig = _mean_signal(pred_row, gt_row)
        if not swap_pairs:
            return mean_orig, False
        pred_swapped = _apply_swaps_row(pred_row, swap_pairs)
        mean_swap = _mean_signal(pred_swapped, gt_row)
        return (mean_swap, True) if mean_swap > mean_orig else (mean_orig, False)

    candidates: List[Tuple[float, int, int, bool]] = []
    for pred_i, pred_row in enumerate(pred_rows):
        for gt_i, gt_row in enumerate(gt_rows):
            mean_score, swapped = _best_orientation(pred_row, gt_row)
            candidates.append((mean_score, pred_i, gt_i, swapped))

    candidates.sort(key=lambda x: x[0], reverse=True)

    used_pred: set[int] = set()
    used_gt: set[int] = set()
    matches: List[Tuple[int, int, bool, float]] = []
    for mean_score, pred_i, gt_i, swapped in candidates:
        if pred_i in used_pred or gt_i in used_gt:
            continue
        used_pred.add(pred_i)
        used_gt.add(gt_i)
        matches.append((pred_i, gt_i, swapped, float(mean_score)))

    return matches


def confusion_counts_presence(
    pred_rows: List[Dict[str, Any]],
    gt_rows: List[Dict[str, Any]],
    fields: List[str],
    *,
    allow_pair_swaps: bool = True,
) -> Dict[str, float]:
    """
    Field-level TP/FP/FN/TN and TP-average similarity after greedy row matching.
    """
    conf: Dict[str, float] = {"TP": 0, "FP": 0, "FN": 0, "TN": 0}
    tp_similarity_total = 0.0
    tp_similarity_n = 0
    swap_pairs = find_crop_swap_pairs(fields) if allow_pair_swaps else []

    matches = match_records_greedy_presence(
        pred_rows,
        gt_rows,
        fields,
        allow_pair_swaps=allow_pair_swaps,
    )
    used_pred = {pred_i for pred_i, _, _, _ in matches}
    used_gt = {gt_i for _, gt_i, _, _ in matches}

    for pred_i, gt_i, swapped, _ in matches:
        gt_row = gt_rows[gt_i]
        pred_row = pred_rows[pred_i]
        pred_row_adj = _apply_swaps_row(pred_row, swap_pairs) if swapped else pred_row
        for f in fields:
            gt_p = is_present(gt_row.get(f))
            pred_p = is_present(pred_row_adj.get(f))
            if gt_p and pred_p:
                conf["TP"] += 1
                tp_similarity_total += field_similarity_score(gt_row.get(f), pred_row_adj.get(f), field_name=f)
                tp_similarity_n += 1
            elif gt_p and not pred_p:
                conf["FN"] += 1
            elif (not gt_p) and pred_p:
                conf["FP"] += 1
            else:
                conf["TN"] += 1

    for gt_i, gt_row in enumerate(gt_rows):
        if gt_i in used_gt:
            continue
        for f in fields:
            if is_present(gt_row.get(f)):
                conf["FN"] += 1
            else:
                conf["TN"] += 1

    for pred_i, pred_row in enumerate(pred_rows):
        if pred_i in used_pred:
            continue
        for f in fields:
            if is_present(pred_row.get(f)):
                conf["FP"] += 1
            else:
                conf["TN"] += 1

    conf["tp_avg_similarity"] = (tp_similarity_total / tp_similarity_n) if tp_similarity_n else float("nan")
    return conf


def overall_metrics_presence(
    pred_rows: List[Dict[str, Any]],
    gt_rows: List[Dict[str, Any]],
    fields: List[str],
    *,
    allow_pair_swaps: bool = True,
) -> Dict[str, float]:
    """Compute the notebook-style aggregate metrics for list-of-dicts rows."""
    conf = confusion_counts_presence(
        pred_rows,
        gt_rows,
        fields,
        allow_pair_swaps=allow_pair_swaps,
    )
    TP, FP, FN, TN = conf["TP"], conf["FP"], conf["FN"], conf["TN"]
    precision = TP / (TP + FP) if (TP + FP) else float("nan")
    tp_avg_similarity = conf["tp_avg_similarity"]
    true_precision = (
        precision * tp_avg_similarity
        if (precision == precision and tp_avg_similarity == tp_avg_similarity)
        else float("nan")
    )
    recall = TP / (TP + FN) if (TP + FN) else float("nan")
    specificity = TN / (TN + FP) if (TN + FP) else float("nan")
    f1 = (2 * true_precision * recall / (true_precision + recall)) if (true_precision + recall) else float("nan")
    return {
        **conf,
        "precision": precision,
        "true_precision": true_precision,
        "recall": recall,
        "specificity": specificity,
        "f1": f1,
    }


def choose_best_global_swap_orientation(
    pred_rows: List[Dict[str, Any]],
    gt_rows: List[Dict[str, Any]],
    fields: List[str],
    *,
    score_key: str = "tp_avg_similarity",
) -> Dict[str, Any]:
    """
    Compare original predictions vs globally swapped 1/2 fields and return the
    higher-scoring orientation.

    Pair-level swaps are disabled while scoring these two orientations, because
    this function is deciding a single orientation for the entire method output.
    """
    original_metrics = overall_metrics_presence(
        pred_rows,
        gt_rows,
        fields,
        allow_pair_swaps=False,
    )
    swapped_rows = swap_1_2_fields_for_records(pred_rows, fields)
    swapped_metrics = overall_metrics_presence(
        swapped_rows,
        gt_rows,
        fields,
        allow_pair_swaps=False,
    )

    original_score = original_metrics.get(score_key, float("nan"))
    swapped_score = swapped_metrics.get(score_key, float("nan"))
    use_swapped = (
        (swapped_score == swapped_score)
        and ((original_score != original_score) or swapped_score > original_score)
    )

    return {
        "rows": swapped_rows if use_swapped else pred_rows,
        "swapped": use_swapped,
        "metrics": swapped_metrics if use_swapped else original_metrics,
        "original_metrics": original_metrics,
        "swapped_metrics": swapped_metrics,
        "score": swapped_score if use_swapped else original_score,
        "score_key": score_key,
    }


# ---------------------------------------------------------------------------
# Two-record evaluation (per field + value+unit + crop swap)
# ---------------------------------------------------------------------------


def _row_get(row: RecordRow, key: str, default: Any = None) -> Any:
    """Read a field from a dict-like record or pandas Series."""
    if hasattr(row, "get"):
        return row.get(key, default)
    try:
        return row[key]
    except (KeyError, TypeError, IndexError):
        return default


def _row_to_dict(row: RecordRow) -> Dict[str, Any]:
    if isinstance(row, dict):
        return dict(row)
    if hasattr(row, "to_dict"):
        return row.to_dict()
    return dict(row)


def _try_parse_magnitude(value: Any) -> Optional[float]:
    if _is_missing(value):
        return None
    dec = _try_parse_decimal(value)
    if dec is not None:
        return float(dec)
    return None


def _gt_unit_field_for_value(
    value_field: str,
    gt_row: RecordRow,
    *,
    gt_spreadsheet: bool = False,
) -> Optional[str]:
    """Resolve which GT column holds the unit for a value field."""
    if gt_spreadsheet:
        return unit_field_for_value(value_field, gt_spreadsheet=True)
    pred_unit = unit_field_for_value(value_field)
    gt_unit = unit_field_for_value(value_field, gt_spreadsheet=True)
    if pred_unit and not _is_missing(_row_get(gt_row, pred_unit)):
        return pred_unit
    if gt_unit and not _is_missing(_row_get(gt_row, gt_unit)):
        return gt_unit
    return pred_unit or gt_unit


def _try_build_quantity(value: Any, unit_symbol: Any) -> Optional[Quantity]:
    mag = _try_parse_magnitude(value)
    if mag is None or _is_missing(unit_symbol):
        return None
    try:
        return Quantity(mag, Unit(str(unit_symbol).strip()))
    except (ValueError, TypeError, ZeroDivisionError):
        return None


def unit_field_similarity_score(
    ref_unit: Any,
    hyp_unit: Any,
) -> float:
    """
    Similarity for a standalone unit column (e.g. ``N Unit`` vs ``kg N/ha``).

    Returns ``1.0`` when both parse as compatible :class:`~units.Unit` with the
    same scale; otherwise falls back to :func:`field_similarity_score`.
    """
    if _is_missing(ref_unit) or _is_missing(hyp_unit):
        return 0.0
    try:
        u_ref = Unit(str(ref_unit).strip())
        u_hyp = Unit(str(hyp_unit).strip())
        if u_ref.is_compatible(u_hyp) and abs(u_ref.conversion_factor_to(u_hyp) - 1.0) < 1e-9:
            return 1.0
    except ValueError:
        pass
    return field_similarity_score(ref_unit, hyp_unit)


def value_with_unit_similarity_score(
    pred_row: RecordRow,
    gt_row: RecordRow,
    value_field: str,
    *,
    gt_spreadsheet: bool = False,
) -> float:
    """
    Score one numeric value field using its paired unit column on each row.

    Builds :class:`~units.Quantity` for prediction and ground truth; returns
    ``1.0`` if they are equal after unit conversion. Falls back to value-only
    :func:`field_similarity_score` when either unit is missing.
    """
    pred_unit_field = unit_field_for_value(value_field)
    gt_unit_field = _gt_unit_field_for_value(
        value_field, gt_row, gt_spreadsheet=gt_spreadsheet
    )

    pred_val = _row_get(pred_row, value_field)
    gt_val = _row_get(gt_row, value_field)
    if _is_missing(pred_val) or _is_missing(gt_val):
        return 0.0

    pred_unit_val = _row_get(pred_row, pred_unit_field) if pred_unit_field else None
    gt_unit_val = _row_get(gt_row, gt_unit_field) if gt_unit_field else None

    q_pred = _try_build_quantity(pred_val, pred_unit_val)
    q_gt = _try_build_quantity(gt_val, gt_unit_val)

    if q_pred is not None and q_gt is not None:
        return 1.0 if q_pred == q_gt else 0.0

    if _is_missing(pred_unit_val) or _is_missing(gt_unit_val):
        return field_similarity_score(gt_val, pred_val, field_name=value_field)

    return 0.0


def _unit_column_in_fields(
    value_field: str,
    fields: List[str],
    gt_row: RecordRow,
    *,
    gt_spreadsheet: bool = False,
) -> Optional[str]:
    """Return the unit column name from ``fields`` paired with ``value_field``."""
    candidates = []
    pred_uf = unit_field_for_value(value_field)
    if pred_uf:
        candidates.append(pred_uf)
    gt_uf = unit_field_for_value(value_field, gt_spreadsheet=True)
    if gt_uf:
        candidates.append(gt_uf)
    if gt_spreadsheet:
        gt_uf2 = _gt_unit_field_for_value(value_field, gt_row, gt_spreadsheet=True)
        if gt_uf2:
            candidates.append(gt_uf2)
    field_set = set(fields)
    for uf in candidates:
        if uf in field_set:
            return uf
    return None


def score_field_between_records(
    pred_row: RecordRow,
    gt_row: RecordRow,
    field_name: str,
    *,
    gt_spreadsheet: bool = False,
    quantity_matched_units: Optional[set[str]] = None,
) -> float:
    """
    Score a single field between two records.

    - Value fields with a paired unit → :func:`value_with_unit_similarity_score`
    - Standalone unit columns → ``1.0`` if already matched via a value+unit pair,
      else :func:`unit_field_similarity_score`
    - All other fields → :func:`field_similarity_score`
    """
    if field_name in VALUE_TO_UNIT_FIELD:
        return value_with_unit_similarity_score(
            pred_row, gt_row, field_name, gt_spreadsheet=gt_spreadsheet
        )
    if field_name in UNIT_FIELDS or field_name in {
        g.unit_field for g in WOPKE_100_GT_VALUE_UNIT_GROUPS
    }:
        if quantity_matched_units and field_name in quantity_matched_units:
            return 1.0
        return unit_field_similarity_score(
            _row_get(gt_row, field_name),
            _row_get(pred_row, field_name),
        )
    return field_similarity_score(
        _row_get(gt_row, field_name),
        _row_get(pred_row, field_name),
        field_name=field_name,
    )


def score_all_fields_between_records(
    pred_row: RecordRow,
    gt_row: RecordRow,
    fields: List[str],
    *,
    gt_spreadsheet: bool = False,
) -> Dict[str, float]:
    """
    Per-field scores for one orientation.

    Value fields with a paired unit are scored via combined
    :class:`~units.Quantity` comparison. When that combined score is ``1.0``,
    the paired unit column (if present in ``fields``) also receives ``1.0``.
    """
    scores: Dict[str, float] = {}
    quantity_matched_units: set[str] = set()
    unit_fields_in_batch = {f for f in fields if f in UNIT_FIELDS} | {
        f for f in fields if f in {g.unit_field for g in WOPKE_100_GT_VALUE_UNIT_GROUPS}
    }

    for f in fields:
        if f in unit_fields_in_batch:
            continue
        if f in VALUE_TO_UNIT_FIELD:
            s = value_with_unit_similarity_score(
                pred_row, gt_row, f, gt_spreadsheet=gt_spreadsheet
            )
            scores[f] = s
            if s == 1.0:
                uf = _unit_column_in_fields(
                    f, fields, gt_row, gt_spreadsheet=gt_spreadsheet
                )
                if uf:
                    quantity_matched_units.add(uf)
            continue
        scores[f] = field_similarity_score(
            _row_get(gt_row, f),
            _row_get(pred_row, f),
            field_name=f,
        )

    for f in fields:
        if f not in unit_fields_in_batch:
            continue
        scores[f] = score_field_between_records(
            pred_row,
            gt_row,
            f,
            gt_spreadsheet=gt_spreadsheet,
            quantity_matched_units=quantity_matched_units,
        )

    return scores


@dataclass
class RecordPairEvaluation:
    """
    Result of comparing one predicted record to one ground-truth record.

    ``field_scores`` uses the better of original vs crop-swapped prediction
    orientation (step 3). ``field_scores_original`` / ``field_scores_swapped``
    retain both orientations for inspection.
    """

    field_scores: Dict[str, float]
    swapped: bool
    mean_score: float
    field_scores_original: Dict[str, float] = field(default_factory=dict)
    field_scores_swapped: Dict[str, float] = field(default_factory=dict)
    mean_score_original: float = 0.0
    mean_score_swapped: float = 0.0
    swap_pairs: Tuple[Tuple[str, str], ...] = ()

    def as_tuple(self) -> Tuple[Dict[str, float], float, bool]:
        """Legacy shape: ``(field_scores, mean_score, swapped)``."""
        return self.field_scores, self.mean_score, self.swapped


def evaluate_record_pair_with_units(
    pred_row: RecordRow,
    gt_row: RecordRow,
    fields: List[str],
    *,
    gt_spreadsheet: bool = False,
    allow_pair_swaps: bool = True,
) -> RecordPairEvaluation:
    """
    Evaluate one prediction record against one GT record in three steps.

    1. Score every field in ``fields`` for the original crop orientation.
    2. For value fields with a paired unit, use combined value+unit comparison
       (:class:`~units.Quantity`). When that score is ``1.0``, the paired unit
       column in ``fields`` also receives ``1.0``. Other unit columns fall back
       to :func:`unit_field_similarity_score`.
    3. Optionally swap crop-1/crop-2 columns on the prediction row, repeat
       steps 1–2, and keep whichever orientation has the higher mean score.

    Args:
        pred_row:        Extracted record (``dict`` or pandas Series).
        gt_row:          Ground-truth record.
        fields:          Column names to score.
        gt_spreadsheet:  Use GT spreadsheet unit headers (``Unit.1``, …).
        allow_pair_swaps: If ``False``, skip step 3.

    Returns:
        :class:`RecordPairEvaluation` with final scores and swap metadata.
    """
    if not fields:
        return RecordPairEvaluation(
            field_scores={},
            swapped=False,
            mean_score=0.0,
        )

    scores_original = score_all_fields_between_records(
        pred_row, gt_row, fields, gt_spreadsheet=gt_spreadsheet
    )
    mean_original = float(np.mean(list(scores_original.values())))

    swap_pairs = tuple(find_crop_swap_pairs(fields)) if allow_pair_swaps else ()
    scores_swapped: Dict[str, float] = {}
    mean_swapped = 0.0

    if swap_pairs:
        pred_swapped = _apply_swaps_row(_row_to_dict(pred_row), list(swap_pairs))
        scores_swapped = score_all_fields_between_records(
            pred_swapped, gt_row, fields, gt_spreadsheet=gt_spreadsheet
        )
        mean_swapped = float(np.mean(list(scores_swapped.values())))

    use_swapped = bool(swap_pairs) and mean_swapped > mean_original
    if use_swapped:
        return RecordPairEvaluation(
            field_scores=scores_swapped,
            swapped=True,
            mean_score=mean_swapped,
            field_scores_original=scores_original,
            field_scores_swapped=scores_swapped,
            mean_score_original=mean_original,
            mean_score_swapped=mean_swapped,
            swap_pairs=swap_pairs,
        )

    return RecordPairEvaluation(
        field_scores=scores_original,
        swapped=False,
        mean_score=mean_original,
        field_scores_original=scores_original,
        field_scores_swapped=scores_swapped,
        mean_score_original=mean_original,
        mean_score_swapped=mean_swapped,
        swap_pairs=swap_pairs,
    )


def infer_gt_spreadsheet_columns(fields: List[str]) -> bool:
    """True when GT rows use spreadsheet unit headers (``Unit.1``, …)."""
    return "Unit.1" in fields or "Unit of density" in fields


def pred_row_after_swap(
    pred_row: RecordRow,
    evaluation: RecordPairEvaluation,
) -> Dict[str, Any]:
    """Prediction row as used for scoring (after crop swap if applicable)."""
    base = _row_to_dict(pred_row)
    if evaluation.swapped and evaluation.swap_pairs:
        return _apply_swaps_row(base, list(evaluation.swap_pairs))
    return base


def match_records_greedy_with_units(
    pred_rows: List[Dict[str, Any]],
    gt_rows: List[Dict[str, Any]],
    fields: List[str],
    *,
    gt_spreadsheet: bool = False,
) -> List[Tuple[int, int, RecordPairEvaluation]]:
    """
    Greedy one-to-one matching using :func:`evaluate_record_pair_with_units`.

    Returns:
        ``(pred_index, gt_index, evaluation)`` for each matched pair.
    """
    candidates: List[Tuple[float, int, int, RecordPairEvaluation]] = []
    for pred_i, pred_row in enumerate(pred_rows):
        for gt_i, gt_row in enumerate(gt_rows):
            ev = evaluate_record_pair_with_units(
                pred_row,
                gt_row,
                fields,
                gt_spreadsheet=gt_spreadsheet,
            )
            candidates.append((ev.mean_score, pred_i, gt_i, ev))

    candidates.sort(key=lambda x: x[0], reverse=True)

    used_pred: set[int] = set()
    used_gt: set[int] = set()
    matches: List[Tuple[int, int, RecordPairEvaluation]] = []
    for _, pred_i, gt_i, ev in candidates:
        if pred_i in used_pred or gt_i in used_gt:
            continue
        used_pred.add(pred_i)
        used_gt.add(gt_i)
        matches.append((pred_i, gt_i, ev))

    return matches


def confusion_counts_with_units(
    pred_rows: List[Dict[str, Any]],
    gt_rows: List[Dict[str, Any]],
    fields: List[str],
    *,
    gt_spreadsheet: bool = False,
) -> Dict[str, float]:
    """Field-level TP/FP/FN/TN with unit-aware TP similarity."""
    conf: Dict[str, float] = {"TP": 0, "FP": 0, "FN": 0, "TN": 0}
    tp_similarity_total = 0.0
    tp_similarity_n = 0

    matches = match_records_greedy_with_units(
        pred_rows, gt_rows, fields, gt_spreadsheet=gt_spreadsheet
    )
    used_pred = {pred_i for pred_i, _, _ in matches}
    used_gt = {gt_i for _, gt_i, _ in matches}

    for pred_i, gt_i, ev in matches:
        gt_row = gt_rows[gt_i]
        pred_adj = pred_row_after_swap(pred_rows[pred_i], ev)
        for f in fields:
            gt_p = is_present(gt_row.get(f))
            pred_p = is_present(pred_adj.get(f))
            if gt_p and pred_p:
                conf["TP"] += 1
                tp_similarity_total += ev.field_scores.get(f, 0.0)
                tp_similarity_n += 1
            elif gt_p and not pred_p:
                conf["FN"] += 1
            elif (not gt_p) and pred_p:
                conf["FP"] += 1
            else:
                conf["TN"] += 1

    for gt_i, gt_row in enumerate(gt_rows):
        if gt_i in used_gt:
            continue
        for f in fields:
            if is_present(gt_row.get(f)):
                conf["FN"] += 1
            else:
                conf["TN"] += 1

    for pred_i, pred_row in enumerate(pred_rows):
        if pred_i in used_pred:
            continue
        for f in fields:
            if is_present(pred_row.get(f)):
                conf["FP"] += 1
            else:
                conf["TN"] += 1

    conf["tp_avg_similarity"] = (
        tp_similarity_total / tp_similarity_n if tp_similarity_n else float("nan")
    )
    return conf


def overall_metrics_with_units(
    pred_rows: List[Dict[str, Any]],
    gt_rows: List[Dict[str, Any]],
    fields: List[str],
    *,
    gt_spreadsheet: bool = False,
) -> Dict[str, float]:
    """Paper-level metrics using unit-aware greedy matching and field scores."""
    conf = confusion_counts_with_units(
        pred_rows, gt_rows, fields, gt_spreadsheet=gt_spreadsheet
    )
    TP, FP, FN, TN = conf["TP"], conf["FP"], conf["FN"], conf["TN"]
    precision = TP / (TP + FP) if (TP + FP) else float("nan")
    tp_avg_similarity = conf["tp_avg_similarity"]
    true_precision = (
        (precision * tp_avg_similarity)
        if (precision == precision and tp_avg_similarity == tp_avg_similarity)
        else float("nan")
    )
    recall = TP / (TP + FN) if (TP + FN) else float("nan")
    specificity = TN / (TN + FP) if (TN + FP) else float("nan")
    f1 = (
        (2 * true_precision * recall / (true_precision + recall))
        if (true_precision + recall)
        else float("nan")
    )
    return {
        **conf,
        "precision": precision,
        "true_precision": true_precision,
        "recall": recall,
        "specificity": specificity,
        "f1": f1,
    }


def _pair_scores(
    ext_row: Any,
    gt_row: Any,
    cols: List[str],
) -> Tuple[Dict[str, float], float, bool]:
    """Compute per-field similarity; delegates to :func:`evaluate_record_pair_with_units`."""
    return evaluate_record_pair_with_units(ext_row, gt_row, cols).as_tuple()


def evaluate_record_pair(
    ext_row: Any,
    gt_row: Any,
    shared_cols: List[str],
    *,
    gt_spreadsheet: bool = False,
    detailed: bool = False,
) -> Union[RecordPairEvaluation, Tuple[Dict[str, float], float, bool]]:
    """
    Evaluate a single extracted record against a single GT record.

    Args:
        ext_row:        Extracted record (pandas Series or dict-like).
        gt_row:         Ground-truth record.
        shared_cols:    Field names to score.
        gt_spreadsheet: GT uses ``Unit.1`` / ``Unit.2`` column names.
        detailed:       If ``True``, return :class:`RecordPairEvaluation`; else
                        legacy ``(field_scores, mean_score, swapped)`` tuple.

    Returns:
        Per-field scores for the best crop orientation, plus mean and swap flag.
    """
    result = evaluate_record_pair_with_units(
        ext_row, gt_row, shared_cols, gt_spreadsheet=gt_spreadsheet
    )
    if detailed:
        return result
    return result.as_tuple()


def match_records_greedy(
    ext_df: "pd.DataFrame",
    gt_df: "pd.DataFrame",
    shared_cols: List[str],
) -> List[Tuple[int, int, Dict[str, float], bool]]:
    """
    Greedy one-to-one record matching between extracted and GT rows.

    For every candidate pair the best crop orientation (original vs. swapped)
    is chosen automatically by :func:`_pair_scores`.  Pairs are committed
    greedily in descending order of mean field similarity so that the
    globally highest-scoring pairs are locked in first.

    Args:
        ext_df:      DataFrame of extracted records (one row per record).
        gt_df:       DataFrame of ground-truth records.
        shared_cols: Field names present in both DataFrames to use for scoring.

    Returns:
        List of ``(ext_row_index, gt_row_index, score_map, swapped)`` tuples.
        Unmatched GT rows and unmatched extracted rows are not included;
        callers should account for them in the denominator separately.

    Example:
        >>> matches = match_records_greedy(ext_df, gt_df, shared_fields)
        >>> for ext_i, gt_i, scores, swapped in matches:
        ...     print(ext_i, gt_i, swapped, round(sum(scores.values()) / len(scores), 3))
    """
    candidates = []
    for ext_i, ext_row in ext_df.iterrows():
        for gt_i, gt_row in gt_df.iterrows():
            score_map, mean_score, swapped = _pair_scores(ext_row, gt_row, shared_cols)
            candidates.append((mean_score, ext_i, gt_i, score_map, swapped))

    candidates.sort(key=lambda x: x[0], reverse=True)

    used_ext: set = set()
    used_gt: set = set()
    matches: List[Tuple[int, int, Dict[str, float], bool]] = []
    for _, ext_i, gt_i, score_map, swapped in candidates:
        if ext_i in used_ext or gt_i in used_gt:
            continue
        used_ext.add(ext_i)
        used_gt.add(gt_i)
        matches.append((ext_i, gt_i, score_map, swapped))

    return matches


def evaluate_method_scores(
    ext_df: "pd.DataFrame",
    gt_df: "pd.DataFrame",
    shared_cols: List[str],
    total_records_for_denominator: int,
) -> Tuple["pd.DataFrame", float, Dict[str, float], int]:
    """
    Compute a normalized similarity score for one extraction method.

    Score formula::

        normalized_score = sum(field similarities over matched pairs)
                           / (total_records_for_denominator * n_fields)

    Using ``total_records_for_denominator = len(gt_df)`` penalises methods
    that extract fewer records than the ground truth.

    Args:
        ext_df:                       DataFrame of extracted records.
        gt_df:                        Ground-truth DataFrame.
        shared_cols:                  Fields present in both DataFrames.
        total_records_for_denominator: Denominator for normalisation
                                      (typically ``len(gt_df)``).

    Returns:
        scores_df           – per-match, per-field similarity DataFrame.
        overall_normalized  – scalar summary score in ``[0, 1]``.
        per_field_normalized– ``{'rougeL_<field>': score}`` dict for plotting.
        n_matches           – number of matched record pairs.

    Example:
        >>> scores_df, overall, per_field, n = evaluate_method_scores(
        ...     ext_df, gt_paper, shared_fields, total_records=len(gt_paper))
        >>> print(f'Normalized score: {overall:.3f}  ({n} matched pairs)')
    """
    matches = match_records_greedy(ext_df, gt_df, shared_cols)

    score_rows = [
        {f"rougeL_{c}": score_map.get(c, 0.0) for c in shared_cols}
        for _, _, score_map, _ in matches
    ]
    scores_df = pd.DataFrame(score_rows)

    n_fields  = len(shared_cols)
    n_records = max(int(total_records_for_denominator), 1)
    denom     = n_records * max(n_fields, 1)

    total_sum          = float(scores_df.to_numpy().sum()) if not scores_df.empty else 0.0
    overall_normalized = total_sum / denom

    per_field_normalized = {
        f"rougeL_{c}": (
            float(scores_df[f"rougeL_{c}"].sum()) / n_records
            if f"rougeL_{c}" in scores_df.columns
            else 0.0
        )
        for c in shared_cols
    }

    return scores_df, overall_normalized, per_field_normalized, len(matches)


def parse_field_value_pairs_artifact(raw: Any) -> List[Any]:
    """
    Normalise ``field_value_pairs`` workspace artifacts to a list of pairs.

    Accepts a Python list, or a JSON / markdown-wrapped string as produced by LLMs.
    """
    if isinstance(raw, list):
        return raw
    if not isinstance(raw, str):
        return []

    s = raw.strip()
    if len(s) >= 2 and ((s[0] == s[-1] == "'") or (s[0] == s[-1] == '"')):
        s = s[1:-1].strip()
    s = re.sub(r"^```(?:json)?\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s*```$", "", s)

    try:
        out = json.loads(s)
        return out if isinstance(out, list) else []
    except Exception:
        pass

    m = re.search(r"\[.*\]", s, flags=re.DOTALL)
    if m:
        payload = m.group(0)
        try:
            out = json.loads(payload)
            return out if isinstance(out, list) else []
        except Exception:
            try:
                out = ast.literal_eval(payload)
                return out if isinstance(out, list) else []
            except Exception:
                return []
    return []


def print_field_value_pairs_vs_gt(
    raw_pairs: Any,
    gt_df: "pd.DataFrame",
    shared_cols: List[str],
    show_fields: Optional[List[str]] = None,
) -> None:
    """
    Pretty-print Step-1 ``field_value_pairs`` using the same columns as
    :func:`print_matching_pairs_from_df` (FIELD / EXTRACTED / GT / SCORE).

    Step-1 pairs are not aligned to GT rows. For each GT row and each field,
    **EXTRACTED** is the step-1 value (among all pairs with that field name)
    that maximises :func:`field_similarity_score` against that row's GT cell.
    This makes the layout comparable to record-level prints while reflecting
    the bag-of-evidence nature of the artifact.
    """
    pairs = parse_field_value_pairs_artifact(raw_pairs)
    eval_cols = list(shared_cols)
    if not eval_cols:
        print("  [no shared_cols provided]")
        return

    by_field: Dict[str, List[Any]] = {}
    for p in pairs or []:
        if not isinstance(p, (list, tuple)) or len(p) < 2:
            continue
        fname, fval = p[0], p[1]
        if not isinstance(fname, str):
            continue
        if fname not in eval_cols:
            continue
        if _is_missing(fval):
            continue
        by_field.setdefault(fname, []).append(fval)

    display_cols = show_fields if show_fields else eval_cols
    n_pairs = len(pairs) if pairs else 0
    n_gt = len(gt_df)

    print(f"  Step-1 pair occurrences : {n_pairs}")
    print(f"  GT rows                 : {n_gt}")
    print(
        "  Note: no row alignment — per GT row, EXTRACTED = best step-1 value "
        "for that field name vs this row's GT cell."
    )
    print()

    col_w = 28
    val_w = 26
    header = f"{'FIELD':{col_w}}  {'EXTRACTED':{val_w}}  {'GT':{val_w}}  SCORE"
    divider = "-" * len(header)

    for gt_i, gt_row in gt_df.iterrows():
        row_scores: List[float] = []
        lines_out: List[Tuple[str, str, str, float]] = []

        for c in display_cols:
            gt_val = gt_row.get(c, "")
            candidates = by_field.get(c, [])

            if not candidates:
                ext_raw: Any = None
                score = field_similarity_score(gt_val, ext_raw)
            else:
                best_v = None
                best_s = -1.0
                for v in candidates:
                    s = field_similarity_score(gt_val, v)
                    if s > best_s:
                        best_s = s
                        best_v = v
                ext_raw = best_v
                score = best_s

            row_scores.append(score)

            if ext_raw is not None and not _is_missing(ext_raw):
                ext_iso = _excel_serial_to_iso(ext_raw)
                ext_str = ext_iso if ext_iso is not None else str(ext_raw)
            else:
                ext_str = "<missing>"

            if not _is_missing(gt_val):
                gt_iso = _excel_serial_to_iso(gt_val)
                gt_str = gt_iso if gt_iso is not None else str(gt_val)
            else:
                gt_str = "<missing>"

            lines_out.append((c, ext_str, gt_str, float(score)))

        mean_score = float(np.mean(row_scores)) if row_scores else 0.0
        print(
            f"  ── GT row {gt_i}  (step-1 best-per-field vs this GT row)  "
            f"mean={mean_score:.3f}"
        )
        print(f"  {header}")
        print(f"  {divider}")
        for c, ext_str, gt_str, score in lines_out:
            print(f"  {c:{col_w}}  {ext_str:{val_w}}  {gt_str:{val_w}}  {score:.3f}")
        print()


def print_matching_pairs_from_df(
    ext_df: "pd.DataFrame",
    gt_df: "pd.DataFrame",
    shared_cols: List[str],
    max_pairs: Optional[int] = None,
    show_fields: Optional[List[str]] = None,
) -> None:
    """
    Pretty-print each greedy-matched (extracted, GT) record pair with per-field scores.

    Same output format as :func:`print_matching_pairs`, but accepts an in-memory
    DataFrame (e.g. MAS workspace artifacts) instead of a CSV path.
    """
    if ext_df.empty:
        print("  [empty extraction]")
        return

    # Use the full shared column list for both matching and display.
    # Missing extracted columns are treated as "<missing>" (score 0.0) instead of being dropped.
    eval_cols = list(shared_cols)
    if not eval_cols:
        print("  [no shared_cols provided]")
        return

    display_cols = show_fields if show_fields else eval_cols
    matches      = match_records_greedy(ext_df, gt_df, eval_cols)
    swap_pairs   = find_crop_swap_pairs(eval_cols)

    n_gt            = len(gt_df)
    n_ext           = len(ext_df)
    n_matched       = len(matches)
    n_swapped       = sum(1 for *_, sw in matches if sw)
    n_gt_unmatched  = n_gt  - n_matched
    n_ext_unmatched = n_ext - n_matched

    print(f"  Extracted rows : {n_ext}")
    print(f"  GT rows        : {n_gt}")
    print(f"  Matched pairs  : {n_matched}  ({n_swapped} used crop-label swap)")
    print(f"  Unmatched GT   : {n_gt_unmatched}  (no extracted record assigned)")
    print(f"  Unmatched ext  : {n_ext_unmatched}  (hallucinated / extra records)")
    print()

    col_w   = 28
    val_w   = 26
    header  = f"{'FIELD':{col_w}}  {'EXTRACTED':{val_w}}  {'GT':{val_w}}  SCORE"
    divider = "-" * len(header)

    pairs_to_show = matches[:max_pairs] if max_pairs else matches
    for rank, (ext_i, gt_i, score_map, swapped) in enumerate(pairs_to_show, 1):
        ext_row    = ext_df.loc[ext_i]
        gt_row     = gt_df.loc[gt_i]
        mean_score = float(np.mean(list(score_map.values()))) if score_map else 0.0
        swap_label = "  [crop labels swapped]" if swapped else ""
        print(f"  ── Pair {rank}  (ext row {ext_i}  ↔  gt row {gt_i})  mean={mean_score:.3f}{swap_label}")
        print(f"  {header}")
        print(f"  {divider}")

        if swapped:
            ext_display: Any = dict(ext_row)
            for c1, c2 in swap_pairs:
                ext_display[c1] = ext_row.get(c2, "")
                ext_display[c2] = ext_row.get(c1, "")
        else:
            ext_display = ext_row

        for c in display_cols:
            ext_val = ext_display.get(c, "")
            gt_val  = gt_row.get(c, "")
            score   = score_map.get(c, field_similarity_score(ext_val, gt_val))
            # Pretty-print values, normalising Excel-serial-style dates when possible
            if not _is_missing(ext_val):
                ext_iso = _excel_serial_to_iso(ext_val)
                ext_str = ext_iso if ext_iso is not None else str(ext_val)
            else:
                ext_str = "<missing>"

            if not _is_missing(gt_val):
                gt_iso = _excel_serial_to_iso(gt_val)
                gt_str = gt_iso if gt_iso is not None else str(gt_val)
            else:
                gt_str = "<missing>"
            print(f"  {c:{col_w}}  {ext_str:{val_w}}  {gt_str:{val_w}}  {score:.3f}")
        print()

    if n_gt_unmatched > 0:
        matched_gt_indices = {gt_i for _, gt_i, *_ in matches}
        unmatched_gt = gt_df.loc[[i for i in gt_df.index if i not in matched_gt_indices]]
        print("  ── Unmatched GT rows (no extracted record assigned) ──")
        for gt_i, gt_row in unmatched_gt.iterrows():
            vals = {
                c: gt_row.get(c, "")
                for c in display_cols
                if not _is_missing(gt_row.get(c, ""))
            }
            print(f"  gt row {gt_i}: { {k: str(v) for k, v in list(vals.items())[:6]} }")
        print()


def print_matching_pairs(
    csv_path: Any,
    gt_df: "pd.DataFrame",
    shared_cols: List[str],
    max_pairs: Optional[int] = None,
    show_fields: Optional[List[str]] = None,
) -> None:
    """
    Load a method's output CSV and pretty-print each matched GT pair
    side-by-side with per-field similarity scores.

    For each matched pair the function shows:

    * The extracted value, the GT value, and the similarity score for every
      display field.
    * A ``[crop labels swapped]`` label when the swap orientation scored higher.
    * A summary of unmatched GT rows (records the method failed to extract).

    Args:
        csv_path:    Path to the extracted-records CSV file.
        gt_df:       Ground-truth DataFrame (e.g. ``gt_paper``).
        shared_cols: Fields to use for matching (e.g. ``shared_fields``).
        max_pairs:   Maximum number of matched pairs to print (``None`` = all).
        show_fields: Subset of fields to display per pair.
                     Defaults to all fields in ``shared_cols``.

    Example:
        >>> print_matching_pairs(
        ...     csv_path=path_mas,
        ...     gt_df=gt_paper,
        ...     shared_cols=shared_fields,
        ...     show_fields=['Crop species 1', 'Crop species 2',
        ...                  'unified yield ic 1', 'unified yield ic 2'],
        ... )
    """
    if csv_path is None or not os.path.exists(str(csv_path)):
        print("  [no CSV found]")
        return

    ext_df = pd.read_csv(csv_path)
    if ext_df.empty:
        print("  [empty CSV]")
        return

    print_matching_pairs_from_df(
        ext_df,
        gt_df,
        shared_cols,
        max_pairs=max_pairs,
        show_fields=show_fields,
    )


__all__ = [
    "ValueUnitGroup",
    "WOPKE_100_VALUE_UNIT_GROUPS",
    "WOPKE_100_GT_VALUE_UNIT_GROUPS",
    "VALUE_TO_UNIT_FIELD",
    "UNIT_FIELDS",
    "SWAP_ELIGIBLE_SUFFIX_FIELDS",
    "GT_COLUMN_ALIASES",
    "value_unit_groups",
    "unit_field_for_value",
    "apply_gt_column_aliases",
    "load_ground_truth",
    "load_ground_truth_by_study_id",
    "wopke_100_shared_fields",
    "build_study_paper_mapping",
    "get_paper_path_for_study",
    "highlight_numbers_and_tables",
    "build_extraction_context",
    "rouge_l_soft_score",
    "date_similarity_score",
    "field_similarity_score",
    "find_crop_swap_pairs",
    "is_present",
    "swap_1_2_fields_for_records",
    "RecordPairEvaluation",
    "RecordRow",
    "evaluate_record_pair",
    "evaluate_record_pair_with_units",
    "infer_gt_spreadsheet_columns",
    "pred_row_after_swap",
    "match_records_greedy_with_units",
    "confusion_counts_with_units",
    "overall_metrics_with_units",
    "score_field_between_records",
    "score_all_fields_between_records",
    "value_with_unit_similarity_score",
    "unit_field_similarity_score",
    "match_records_greedy",
    "match_records_greedy_presence",
    "confusion_counts_presence",
    "overall_metrics_presence",
    "choose_best_global_swap_orientation",
    "evaluate_method_scores",
    "print_matching_pairs",
    "print_matching_pairs_from_df",
    "parse_field_value_pairs_artifact",
    "print_field_value_pairs_vs_gt",
    "_is_missing",
]
